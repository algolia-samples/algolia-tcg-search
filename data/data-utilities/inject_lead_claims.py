#!/usr/bin/env python3
"""
inject_lead_claims.py — Match Pokemon card claims to a conference leads spreadsheet.

Reads a claims CSV (exported from Supabase), fuzzy-matches claimant names against
a sheet in a conference leads XLSX, and writes the highest-value card name to a
designated column. Safe by default: runs as a dry run unless --write is passed.

Usage:
  python inject_lead_claims.py XLSX CLAIMS_CSV             # dry run
  python inject_lead_claims.py XLSX CLAIMS_CSV --write     # overwrite XLSX in place
  python inject_lead_claims.py XLSX CLAIMS_CSV --write --out updated.xlsx
  python inject_lead_claims.py XLSX CLAIMS_CSV --sheet "Sheet Name" --col-pokemon "Card Won"

Defaults (matching the Microsoft Build 2026 booth leads sheet):
  --sheet      "Booth Leads - TO UPLOAD"
  --col-first  "First Name"
  --col-last   "Last Name"
  --col-pokemon "Pokemon Card"

# TODO: To make this a permanent pipeline tool:
# - Pull claims directly from Supabase (supabase_claims_table.sql) instead of
#   requiring a manual CSV export. The claims table has claimer_first_name,
#   claimer_last_name, card_value, and pokemon_name — same fields used here.
# - Accept --event-id to filter claims to a single event (claims table has event_id).
# - Consider writing back to Supabase (a lead_claims join table) instead of XLSX,
#   and letting the frontend read from there.
"""

import argparse
import csv
import difflib
from pathlib import Path

import openpyxl

FIRST_THRESHOLD = 0.7  # looser — first names vary (Rob/Robert, Tru/Truman)
LAST_THRESHOLD = 0.8   # tighter — last names should be a strong signal


def normalize(s) -> str:
    return str(s or "").strip().lower()


def similarity(a: str, b: str) -> float:
    return difflib.SequenceMatcher(None, a, b).ratio()


def load_best_claims(csv_path: Path) -> dict[tuple[str, str], str]:
    """Return {(first, last): pokemon_name} keeping the highest-value card per person."""
    best: dict[tuple[str, str], tuple[float, str]] = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            first = normalize(row["claimer_first_name"])
            last = normalize(row["claimer_last_name"])
            if not first and not last:
                continue
            key = (first, last)
            value = float(row["card_value"] or 0)
            if key not in best or value > best[key][0]:
                best[key] = (value, row["pokemon_name"])
    return {k: v[1] for k, v in best.items()}


def find_match(
    first: str, last: str, claims: dict[tuple[str, str], str]
) -> tuple[str, str] | None:
    """Exact match first, then fuzzy requiring BOTH first and last to be similar.

    Scoring both name parts independently prevents false positives where only
    a first name or only a last name happens to match.
    """
    if (first, last) in claims:
        return (first, last)

    best_score = 0.0
    best_key = None
    for cf, cl in claims:
        fs = similarity(first, cf)
        ls = similarity(last, cl)
        if fs >= FIRST_THRESHOLD and ls >= LAST_THRESHOLD:
            score = (fs + ls) / 2
            if score > best_score:
                best_score = score
                best_key = (cf, cl)
    return best_key


def find_col(header_row, label: str) -> int:
    """Return 1-based column index for a header label (case-insensitive, stripped)."""
    for i, cell in enumerate(header_row, start=1):
        if normalize(cell.value) == normalize(label):
            return i
    raise ValueError(f"Column '{label}' not found in sheet header row.")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("xlsx", type=Path, help="Path to the conference leads XLSX.")
    parser.add_argument("claims_csv", type=Path, help="Path to the claims CSV export.")
    parser.add_argument("--sheet", default="Booth Leads - TO UPLOAD", help="Sheet name to update.")
    parser.add_argument("--col-first", default="First Name", help="Header for the first-name column.")
    parser.add_argument("--col-last", default="Last Name", help="Header for the last-name column.")
    parser.add_argument("--col-pokemon", default="Pokemon Card", help="Header for the card column to write.")
    parser.add_argument("--write", action="store_true", help="Apply updates. Without this, only a dry run is performed.")
    parser.add_argument("--out", type=Path, help="Output path (default: overwrite input).")
    args = parser.parse_args()

    dry_run = not args.write
    out_path = args.out or args.xlsx

    if dry_run:
        print("DRY RUN — no file will be changed. Pass --write to apply.\n")

    claims = load_best_claims(args.claims_csv)
    print(f"Loaded {len(claims)} unique claimants from {args.claims_csv.name}.")

    wb = openpyxl.load_workbook(args.xlsx)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {args.sheet!r} not found in {args.xlsx.name}. "
            f"Available: {', '.join(wb.sheetnames)}"
        )
    ws = wb[args.sheet]

    header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
    col_first = find_col(header_row, args.col_first)
    col_last = find_col(header_row, args.col_last)
    col_pokemon = find_col(header_row, args.col_pokemon)

    updates = []
    unmatched = []

    for row in ws.iter_rows(min_row=2):
        first = normalize(row[col_first - 1].value)
        last = normalize(row[col_last - 1].value)
        if not first and not last:
            continue

        matched_key = find_match(first, last, claims)
        if matched_key is None:
            unmatched.append(f"{first} {last}")
            continue

        pokemon = claims[matched_key]
        if matched_key == (first, last):
            match_type = "exact"
        else:
            match_type = f"fuzzy via '{matched_key[0]} {matched_key[1]}'"

        cell = row[col_pokemon - 1]
        if cell.value == pokemon:
            continue

        updates.append((cell, pokemon, f"{first} {last}", match_type))

    print(f"\n{len(updates)} rows to update, {len(unmatched)} with no card claim.")

    if updates:
        print("\nPlanned updates:")
        for _, pokemon, name, match_type in updates:
            print(f"  {name:35s}  ->  {pokemon}  ({match_type})")

    if dry_run:
        print("\nDry run complete. Run with --write to apply.")
        return

    if not updates:
        print("\nNothing to update.")
        return

    for cell, pokemon, _, _ in updates:
        cell.value = pokemon

    wb.save(out_path)
    print(f"\nSaved {len(updates)} updates to {out_path}")


if __name__ == "__main__":
    main()
