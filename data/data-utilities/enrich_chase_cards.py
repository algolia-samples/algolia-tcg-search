#!/usr/bin/env python3
"""
Enrich chase cards in Algolia with image URLs and chase_rank.

Reads the Chase Cards sheet from the XLSX export, extracts hyperlinks,
and updates matching Algolia records.

Run this after a full ingest.
"""

import os
import re
import sys
import requests
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv
import openpyxl
from algoliasearch.search.client import SearchClientSync

# Load environment variables
load_dotenv(Path(__file__).parent.parent / ".env")

# Configuration
ALGOLIA_APP_ID = os.getenv("ALGOLIA_APP_ID")
ALGOLIA_API_KEY = os.getenv("ALGOLIA_API_KEY")
ALGOLIA_EVENT_ID = os.getenv("ALGOLIA_EVENT_ID")
if not ALGOLIA_EVENT_ID:
    print("ERROR: ALGOLIA_EVENT_ID is not set. Set it in data/.env or export it before running.")
    sys.exit(1)
ALGOLIA_INDEX_NAME = f"tcg_cards_{ALGOLIA_EVENT_ID}"
DATA_DIR = Path(__file__).parent.parent / "data-files" / ALGOLIA_EVENT_ID
XLSX_FILE = DATA_DIR / "TCG Search Website - Raw List.xlsx"
CHASE_SHEET = "Chase Cards"
TCGDEX_BASE_URL = "https://api.tcgdex.net/v2/en"


def tcgplayer_image_urls(product_url: str) -> tuple[str, str]:
    """
    Extract product ID from a TCGPlayer URL and return (image_small, image_large).
    e.g. https://www.tcgplayer.com/product/632994/... ->
         ('https://tcgplayer-cdn.tcgplayer.com/product/632994_200w.jpg',
          'https://tcgplayer-cdn.tcgplayer.com/product/632994_in_1000x1000.jpg')
    """
    match = re.search(r"/product/(\d+)/", product_url)
    if not match:
        raise ValueError(f"Could not extract product ID from URL: {product_url}")
    product_id = match.group(1)
    base = f"https://tcgplayer-cdn.tcgplayer.com/product/{product_id}"
    return f"{base}_200w.jpg", f"{base}_in_1000x1000.jpg"


def normalize_set_name(set_name: str) -> str:
    """Strip trailing (NNN) count from set names, e.g. 'Journey Together (159)' -> 'Journey Together'."""
    return re.sub(r"\s*\(\d+\)\s*$", "", set_name.strip())


def fetch_set_id_map() -> dict:
    """
    Fetch all sets from TCGdex and return a map of lowercase name -> set_id.
    Includes both full names and names without series prefix for flexible matching.
    """
    print("Fetching TCGdex set list...")
    response = requests.get(f"{TCGDEX_BASE_URL}/sets", timeout=10)
    response.raise_for_status()
    sets = response.json()

    set_id_map = {}
    for s in sets:
        name = s.get("name", "").lower()
        set_id_map[name] = s["id"]
        if ":" in name:
            without_prefix = name.split(":", 1)[1].strip()
            set_id_map.setdefault(without_prefix, s["id"])

    print(f"  Loaded {len(sets)} sets from TCGdex")
    return set_id_map


def resolve_set_id(set_name_raw: str, set_id_map: dict) -> Optional[str]:
    """Resolve a set name from the chase file to a TCGdex set_id."""
    normalized = normalize_set_name(set_name_raw).lower()

    if normalized in set_id_map:
        return set_id_map[normalized]

    if ":" in normalized:
        without_prefix = normalized.split(":", 1)[1].strip()
        if without_prefix in set_id_map:
            return set_id_map[without_prefix]

    return None


def load_chase_rows() -> list[dict]:
    """Load and parse rows from the Chase Cards sheet, including hyperlink URLs."""
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb[CHASE_SHEET]

    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]
    name_col = headers.index("Pokemon Name")
    number_col = headers.index("Number")
    set_col = headers.index("Set")
    link_col = headers.index("Link to picture")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        name = row[name_col].value
        if not name or str(name).strip().lower() == "nan":
            continue
        link_cell = row[link_col]
        url = link_cell.hyperlink.target if link_cell.hyperlink else None
        rows.append({
            "pokemon_name": str(name).strip(),
            "raw_number": str(row[number_col].value).strip() if row[number_col].value else "",
            "set_name_raw": str(row[set_col].value).strip() if row[set_col].value else "",
            "image_url": url,
        })
    return rows


def validate_chase_cards(client: SearchClientSync, set_id_map: dict) -> list[dict]:
    """
    For each chase card row, construct the expected objectID and verify it exists
    in the Algolia index. Returns a list of resolved records with objectID populated.
    """
    rows = load_chase_rows()
    print(f"Loaded {len(rows)} chase card rows from XLSX\n")

    # Build objectID for each row
    resolved = []
    unresolvable = []

    for row in rows:
        raw = row["raw_number"].split("/")[0].strip()
        # Preserve both formats: ingest.py may store with or without leading zeros
        # depending on whether pandas read the column as string or float
        card_number = raw
        card_number_stripped = str(int(raw))
        set_id = resolve_set_id(row["set_name_raw"], set_id_map)

        if not set_id:
            unresolvable.append(row)
            print(f"  ✗ Could not resolve set: '{row['set_name_raw']}' ({row['pokemon_name']})")
            continue

        resolved.append({**row, "card_number": card_number, "card_number_stripped": card_number_stripped, "set_id": set_id})

    if unresolvable:
        print(f"\n  {len(unresolvable)} rows could not be resolved to a set ID\n")

    # Search Algolia for each card using set_name + number filters
    print(f"Verifying {len(resolved)} cards in Algolia...")
    found = []
    missing = []

    for chase_rank, record in enumerate(resolved, start=1):
        record["chase_rank"] = chase_rank
        set_name = normalize_set_name(record["set_name_raw"])
        set_filter = f'set_name:"{set_name}"'
        hit = None

        # Try original number (may have leading zeros), then stripped
        for query in [record["card_number"], record["card_number_stripped"]]:
            result = client.search_single_index(
                index_name=ALGOLIA_INDEX_NAME,
                search_params={"query": query, "filters": set_filter, "hitsPerPage": 1},
            )
            if result.hits:
                hit = result.hits[0]
                break

        if hit:
            record["objectID"] = hit.object_id
            found.append(record)
            print(f"  ✓ {record['objectID']}  ({record['pokemon_name']})")
        else:
            missing.append(record)
            print(f"  ✗ NOT FOUND: set='{set_name}' number={record['card_number']}  ({record['pokemon_name']})")

    print(f"\nResults: {len(found)} found, {len(missing)} missing, {len(unresolvable)} unresolvable")
    return found


def apply_chase_updates(client: SearchClientSync, found: list[dict]):
    """Partial-update each found record with image URLs and chase_rank."""
    updates = []
    for record in found:
        try:
            image_small, image_large = tcgplayer_image_urls(record["image_url"])
        except ValueError as e:
            print(f"  ⚠ Skipping {record['objectID']}: {e}")
            continue
        updates.append({
            "objectID": record["objectID"],
            "image_small": image_small,
            "image_large": image_large,
            "chase_rank": record["chase_rank"],
        })

    print(f"\nApplying {len(updates)} partial updates to Algolia...")
    try:
        client.partial_update_objects(index_name=ALGOLIA_INDEX_NAME, objects=updates)
        print(f"✓ Successfully updated {len(updates)} chase card records")
    except Exception as e:
        print(f"✗ Error updating Algolia: {e}")


def main():
    if not ALGOLIA_APP_ID or not ALGOLIA_API_KEY:
        print("ERROR: Missing Algolia credentials in .env file")
        return

    if not XLSX_FILE.exists():
        print(f"ERROR: XLSX file not found: {XLSX_FILE}")
        return

    set_id_map = fetch_set_id_map()
    client = SearchClientSync(ALGOLIA_APP_ID, ALGOLIA_API_KEY)

    found = validate_chase_cards(client, set_id_map)
    if found:
        apply_chase_updates(client, found)


if __name__ == "__main__":
    main()
